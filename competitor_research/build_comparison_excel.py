# -*- coding: utf-8 -*-
"""Create Excel (.xlsx) - one sheet per destination from comparison_by_destination.json"""
import json
from pathlib import Path

BASE = Path(__file__).resolve().parent
COMPARISON_PATH = BASE / "comparison_by_destination.json"
FULL_SCAN_PATH = BASE / "full_scan_results.json"
XLSX_PATH = BASE / "השוואת_טיולים_לפי_יעד.xlsx"

# Excel sheet names max 31 chars; map long names to short
SHEET_NAMES = {
    "יפן": "יפן",
    "פיורדים": "פיורדים",
    "מזרח הרחוק": "מזרח הרחוק",
    "דנובה": "דנובה",
    "דואורו": "דואורו",
    "ים תיכון": "ים תיכון",
    "אלסקה": "אלסקה",
    "שייט נהרות – רון, סיין, דורדון, ריין": "שייט נהרות",
    "ים בלטי": "ים בלטי",
}


def load_comparison():
    with open(COMPARISON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_full_scan():
    with open(FULL_SCAN_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _str(val):
    if val is None:
        return ""
    return str(val)


def build_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("נדרש: pip install openpyxl")
        return False

    data = load_comparison()
    rtl = Alignment(horizontal="right", vertical="center", reading_order=2)
    bold = Font(bold=True)

    wb = openpyxl.Workbook()
    # Remove default sheet, we'll create one per destination
    wb.remove(wb.active)

    for dest_key, dest_data in data["destinations"].items():
        sheet_title = SHEET_NAMES.get(dest_key, dest_key[:31])
        ws = wb.create_sheet(sheet_title)
        ws.sheet_view.rightToLeft = True

        # תרבותו
        ws.append(["תרבותו"])
        ws.cell(ws.max_row, 1).font = bold
        headers_t = ["שם הטיול", "ימים", "מחיר", "תאריך"]
        ws.append(headers_t)
        for c in range(1, len(headers_t) + 1):
            ws.cell(ws.max_row, c).font = bold
            ws.cell(ws.max_row, c).alignment = rtl
        for t in dest_data["tarbutu"]:
            row = [t.get("name", ""), _str(t.get("days")), _str(t.get("price")), _str(t.get("date"))]
            if t.get("destination"):
                row.append(_str(t.get("destination")))
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row:
                c.alignment = rtl

        # מתחרים
        ws.append([])
        ws.append(["מתחרים"])
        ws.cell(ws.max_row, 1).font = bold
        headers_c = ["חברה", "שם הטיול", "ימים", "מחיר", "תאריך", "הערות"]
        ws.append(headers_c)
        for c in range(1, len(headers_c) + 1):
            ws.cell(ws.max_row, c).font = bold
            ws.cell(ws.max_row, c).alignment = rtl
        for t in dest_data["competitors"]:
            note = t.get("note") or ""
            if t.get("guaranteed"):
                note = "מובטח" + ("; " + note if note else "")
            ws.append([
                _str(t.get("company")),
                _str(t.get("name")),
                _str(t.get("days")),
                _str(t.get("price")),
                _str(t.get("date")),
                note,
            ])
        for row in ws.iter_rows(min_row=ws.max_row - len(dest_data["competitors"]), max_row=ws.max_row):
            for c in row:
                c.alignment = rtl

    # Sheet: סיכום תרבותו (all Tarbutu trips)
    full = load_full_scan()
    ws_summary = wb.create_sheet("סיכום תרבותו")
    ws_summary.sheet_view.rightToLeft = True
    ws_summary.append(["סיכום כל הטיולים של תרבותו", "", "", "", ""])
    ws_summary.cell(1, 1).font = bold
    ws_summary.append([])

    ws_summary.append(["טיולי תרבות יבשתיים"])
    ws_summary.cell(ws_summary.max_row, 1).font = bold
    ws_summary.append(["שם הטיול", "ימים", "תאריך", "יעד", "מחיר"])
    for r in ws_summary.iter_rows(min_row=ws_summary.max_row, max_row=ws_summary.max_row):
        for c in r:
            c.font, c.alignment = bold, rtl
    for t in full["tarbutu"]["land_tours"]:
        ws_summary.append([t["name"], _str(t.get("days")), _str(t.get("date")), _str(t.get("destination")), "לפי פנייה"])
    for row in ws_summary.iter_rows(min_row=ws_summary.max_row - len(full["tarbutu"]["land_tours"]), max_row=ws_summary.max_row):
        for c in row:
            c.alignment = rtl

    ws_summary.append([])
    ws_summary.append(["קרוזי ים"])
    ws_summary.cell(ws_summary.max_row, 1).font = bold
    ws_summary.append(["שם הטיול", "ימים", "תאריך", "יעד/הערות", "מחיר"])
    for r in ws_summary.iter_rows(min_row=ws_summary.max_row, max_row=ws_summary.max_row):
        for c in r:
            c.font, c.alignment = bold, rtl
    for t in full["tarbutu"]["sea_cruises"]:
        dest = t.get("destination") or t.get("category") or ""
        ws_summary.append([t["name"], _str(t.get("days")), _str(t.get("date")), dest, "לפי פנייה"])
    for row in ws_summary.iter_rows(min_row=ws_summary.max_row - len(full["tarbutu"]["sea_cruises"]), max_row=ws_summary.max_row):
        for c in row:
            c.alignment = rtl

    ws_summary.append([])
    ws_summary.append(["שייט נהרות"])
    ws_summary.cell(ws_summary.max_row, 1).font = bold
    ws_summary.append(["שם הטיול", "ימים", "תאריך", "ספינה/הערות", "מחיר"])
    for r in ws_summary.iter_rows(min_row=ws_summary.max_row, max_row=ws_summary.max_row):
        for c in r:
            c.font, c.alignment = bold, rtl
    for t in full["tarbutu"]["river_cruises"]:
        ship = t.get("ship") or t.get("ships") or t.get("note") or ""
        ws_summary.append([t["name"], _str(t.get("days")), _str(t.get("date")), ship, "לפי פנייה"])
    for row in ws_summary.iter_rows(min_row=ws_summary.max_row - len(full["tarbutu"]["river_cruises"]), max_row=ws_summary.max_row):
        for c in row:
            c.alignment = rtl

    wb.save(XLSX_PATH)
    print("נשמר:", XLSX_PATH)
    return True


if __name__ == "__main__":
    build_excel()
